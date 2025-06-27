import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_prices = 0.9 * cell.value
        corrected_prices_cell = sheet.cell(row=row, column=4)
        corrected_prices_cell.value = corrected_prices

    values = Reference(sheet,
              min_row=2, max_row=sheet.max_row,
              max_col=4, min_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'E2')
    wb.save(filename)


